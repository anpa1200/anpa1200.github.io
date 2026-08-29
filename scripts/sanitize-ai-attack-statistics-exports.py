#!/usr/bin/env python3
"""Create rights-safe, formula-safe public AI-attack research exports."""

from __future__ import annotations

import csv
import json
import os
from pathlib import Path
import re
import sqlite3
import tempfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "ai-attack-statistics" / "data"
CSV_EXPORTS = (
    "publications.csv",
    "tags_long.csv",
    "metrics_long.csv",
    "iocs_long.csv",
    "quality.csv",
    "tag_dictionary.csv",
)
FORMULA_PREFIXES = ("=", "+", "-", "@")
CSV_DROP_COLUMNS = {
    "publications.csv": {"local_files"},
    "tags_long.csv": {"evidence_quote"},
    "metrics_long.csv": {"evidence_quote"},
    "iocs_long.csv": {"evidence_quote"},
}
WORKBOOK_DROP_COLUMNS = {
    "Publications": {"local_files"},
    "Tags": {"evidence_quote"},
    "Metrics": {"evidence_quote"},
    "IOCs": {"evidence_quote"},
}


def literal(value: object) -> object:
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def defang_domain(value: object) -> object:
    """Render a domain candidate inert without changing its analytical identity."""
    if not isinstance(value, str):
        return value
    return re.sub(r"(?<!\[)\.(?!\])", "[.]", value)


def sanitize_csv(path: Path) -> tuple[int, int, int]:
    changed = 0
    defanged = 0
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        rows = list(csv.reader(source))
    if not rows:
        return 0, 0, 0
    dropped = CSV_DROP_COLUMNS.get(path.name, set())
    keep = [index for index, name in enumerate(rows[0]) if name not in dropped]
    dropped_count = len(rows[0]) - len(keep)
    rows = [[row[index] if index < len(row) else "" for index in keep] for row in rows]
    if path.name == "iocs_long.csv":
        ioc_type_index = rows[0].index("ioc_type")
        value_index = rows[0].index("value")
        for row in rows[1:]:
            if row[ioc_type_index] != "defanged_domain":
                continue
            safe = defang_domain(row[value_index])
            if safe != row[value_index]:
                row[value_index] = str(safe)
                defanged += 1
    for row in rows:
        for index, value in enumerate(row):
            safe = literal(value)
            if safe != value:
                row[index] = str(safe)
                changed += 1
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        newline="",
        dir=path.parent,
        delete=False,
    ) as target:
        writer = csv.writer(target, lineterminator="\n")
        writer.writerows(rows)
        temporary = Path(target.name)
    os.replace(temporary, path)
    return changed, dropped_count, defanged


def sanitize_tsv(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        rows = list(csv.reader(source, delimiter="\t"))
    if not rows:
        return 0
    keep = [index for index, name in enumerate(rows[0]) if name != "file"]
    dropped_count = len(rows[0]) - len(keep)
    rows = [[row[index] if index < len(row) else "" for index in keep] for row in rows]
    if "confirmed_companion_group" not in rows[0]:
        rows[0].append("confirmed_companion_group")
        for row in rows[1:]:
            row.append("")

    with (DATA / "publications.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as source:
        publications = list(csv.DictReader(source))
    companion_groups: dict[str, str] = {}
    for publication in publications:
        group = publication.get("duplicate_group", "").strip()
        if not group:
            continue
        source_ids = [publication.get("primary_source_id", "")]
        source_ids.extend(publication.get("companion_source_ids", "").split("|"))
        for source_id in source_ids:
            if source_id.strip():
                companion_groups[source_id.strip()] = group

    source_id_index = rows[0].index("id")
    companion_group_index = rows[0].index("confirmed_companion_group")
    review_status_index = (
        rows[0].index("review_status") if "review_status" in rows[0] else None
    )
    for row in rows[1:]:
        row.extend("" for _ in range(len(rows[0]) - len(row)))
        row[companion_group_index] = companion_groups.get(row[source_id_index], "none")
        if row[companion_group_index] != "none" and review_status_index is not None:
            row[review_status_index] = "confirmed_companion_format_duplicate"
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        newline="",
        dir=path.parent,
        delete=False,
    ) as target:
        writer = csv.writer(target, delimiter="\t", lineterminator="\n")
        writer.writerows(rows)
        temporary = Path(target.name)
    os.replace(temporary, path)
    return dropped_count


def sanitize_workbook(path: Path) -> tuple[int, int, int, int]:
    workbook = load_workbook(path, data_only=False)
    changed = 0
    dropped_count = 0
    defanged = 0
    filter_repairs = 0
    for worksheet in workbook.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        drop = WORKBOOK_DROP_COLUMNS.get(worksheet.title, set())
        for column in sorted(
            (index for index, name in enumerate(headers, start=1) if name in drop),
            reverse=True,
        ):
            worksheet.delete_cols(column)
            dropped_count += 1
        headers = [cell.value for cell in worksheet[1]]
        if worksheet.title == "IOCs":
            ioc_type_column = headers.index("ioc_type") + 1
            value_column = headers.index("value") + 1
            for row in range(2, worksheet.max_row + 1):
                if worksheet.cell(row, ioc_type_column).value != "defanged_domain":
                    continue
                cell = worksheet.cell(row, value_column)
                safe = defang_domain(cell.value)
                if safe != cell.value:
                    cell.value = safe
                    cell.data_type = "s"
                    defanged += 1
        for row in worksheet.iter_rows():
            for cell in row:
                safe = literal(cell.value)
                if safe != cell.value:
                    cell.value = safe
                    cell.data_type = "s"
                    changed += 1
        if (
            worksheet.auto_filter.ref
            and worksheet.auto_filter.ref != worksheet.dimensions
        ):
            worksheet.auto_filter.ref = worksheet.dimensions
            filter_repairs += 1
    if changed or dropped_count or defanged or filter_repairs:
        workbook.save(path)
    return changed, dropped_count, defanged, filter_repairs


def sanitize_sqlite(path: Path) -> tuple[int, int, int]:
    connection = sqlite3.connect(path)
    dropped_tables = 0
    dropped_columns = 0
    defanged = 0
    try:
        if connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'source_records'"
        ).fetchone():
            connection.execute("DROP TABLE source_records")
            dropped_tables += 1
        for table, column in (
            ("publications", "local_files"),
            ("tags", "evidence_quote"),
            ("metrics", "evidence_quote"),
            ("iocs", "evidence_quote"),
        ):
            columns = {
                row[1] for row in connection.execute(f'PRAGMA table_info("{table}")')
            }
            if column in columns:
                connection.execute(f'ALTER TABLE "{table}" DROP COLUMN "{column}"')
                dropped_columns += 1
        for rowid, value in connection.execute(
            "SELECT rowid, value FROM iocs WHERE ioc_type = 'defanged_domain'"
        ):
            safe = defang_domain(value)
            if safe != value:
                connection.execute(
                    "UPDATE iocs SET value = ? WHERE rowid = ?", (safe, rowid)
                )
                defanged += 1
        if dropped_tables or dropped_columns or defanged:
            connection.commit()
            connection.execute("VACUUM")
    finally:
        connection.close()
    return dropped_tables, dropped_columns, defanged


def update_summary(path: Path) -> dict[str, int]:
    with (DATA / "publications.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as source:
        publications = list(csv.DictReader(source))
    eligible_ids = {
        row["publication_id"]
        for row in publications
        if row["analysis_inclusion"] == "include_with_manual_validation"
    }

    with (DATA / "tags_long.csv").open("r", encoding="utf-8-sig", newline="") as source:
        eligible_tags = [
            row
            for row in csv.DictReader(source)
            if row["publication_id"] in eligible_ids
        ]
    with (DATA / "metrics_long.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as source:
        eligible_metrics = [
            row
            for row in csv.DictReader(source)
            if row["publication_id"] in eligible_ids
        ]
    with (DATA / "iocs_long.csv").open("r", encoding="utf-8-sig", newline="") as source:
        eligible_iocs = [
            row
            for row in csv.DictReader(source)
            if row["publication_id"] in eligible_ids
        ]

    values = {
        "eligible_iocs": len(eligible_iocs),
        "eligible_metrics": len(eligible_metrics),
        "eligible_publications_with_iocs": len(
            {row["publication_id"] for row in eligible_iocs}
        ),
        "eligible_publications_with_metrics": len(
            {row["publication_id"] for row in eligible_metrics}
        ),
        "eligible_tag_types": len({row["tag_type"] for row in eligible_tags}),
        "eligible_tags": len(eligible_tags),
        "eligible_unique_tag_values": len(
            {(row["tag_type"], row["normalized_value"]) for row in eligible_tags}
        ),
    }
    summary = json.loads(path.read_text(encoding="utf-8"))
    summary.update(values)
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        dir=path.parent,
        delete=False,
    ) as target:
        json.dump(summary, target, indent=2, sort_keys=True)
        target.write("\n")
        temporary = Path(target.name)
    os.replace(temporary, path)
    return values


def main() -> None:
    csv_results = [sanitize_csv(DATA / name) for name in CSV_EXPORTS]
    csv_changes = sum(result[0] for result in csv_results)
    csv_columns = sum(result[1] for result in csv_results)
    csv_defanged = sum(result[2] for result in csv_results)
    tsv_columns = sanitize_tsv(DATA / "source-uniqueness-report.tsv")
    workbook_changes, workbook_columns, workbook_defanged, workbook_filters = (
        sanitize_workbook(DATA / "ai_attack_statistics.xlsx")
    )
    sqlite_tables, sqlite_columns, sqlite_defanged = sanitize_sqlite(
        DATA / "ai_attack_statistics.sqlite"
    )
    eligible_summary = update_summary(DATA / "summary.json")
    print(
        "Sanitized AI-attack exports: "
        f"{csv_changes} CSV cells and {workbook_changes} workbook cells literalized; "
        f"removed {csv_columns + tsv_columns} delimited, {workbook_columns} workbook, "
        f"and {sqlite_columns} SQLite evidence/path columns plus {sqlite_tables} private table; "
        f"defanged {csv_defanged} CSV, {workbook_defanged} workbook, and "
        f"{sqlite_defanged} SQLite domain candidates; repaired {workbook_filters} workbook filters; "
        f"maintained {len(eligible_summary)} eligible-summary fields."
    )


if __name__ == "__main__":
    main()
